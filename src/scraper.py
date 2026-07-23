from selenium import webdriver 
from selenium.webdriver.common.by import By
import time 
import pandas as pd
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

driver = webdriver.Chrome()

base_url = "https://aqarmap.com.eg/ar/for-sale/apartment/?page="


properties = []

MAX_PAGES = 11
for page in range(1, MAX_PAGES + 1):

    print("Opening page:", page)

    driver.get(base_url + str(page))

    WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.TAG_NAME, "article"))
    )

    cards = driver.find_elements(By.TAG_NAME,"article")

    print("Cards found:", len(cards))

    for card in cards :
        
        title = card.find_element(By.TAG_NAME, "h2").text
        price = card.find_element(By.TAG_NAME, "data").get_attribute("value")
        links = card.find_elements(By.TAG_NAME, "a")

        location = links[-2].text

        details = card.find_elements(By.TAG_NAME,"span")


        area = ""
        bedrooms = ""
        bathrooms = ""

        for detail in details :
            text = detail.text

            if "م²" in text:
                area = text

            elif text.isdigit():
                if bedrooms == "":
                    bedrooms = text
                else:
                    bathrooms = text

        properties.append({
            "title": title,
            "price": price,
            "location": location,
            "area": area,
            "bedrooms": bedrooms,
            "bathrooms": bathrooms
        })

df = pd.DataFrame(properties)

file_path = r"G:\test wep scraping\real-estate-scraper\data\real_estate.xlsx"

df.to_excel(file_path, index=False)

workbook = load_workbook(file_path)
sheet = workbook.active


# تنسيق العناوين
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)

for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")


# إضافة حدود
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row in sheet.iter_rows():
    for cell in row:
        cell.border = thin_border


# ضبط عرض الأعمدة
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    sheet.column_dimensions[column_letter].width = max_length + 3


workbook.save(file_path)

driver.quit()

print(f"Scraped {len(df)} properties.")
print(f"Data saved successfully: {file_path}")